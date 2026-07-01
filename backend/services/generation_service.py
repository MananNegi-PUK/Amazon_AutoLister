import os
import openpyxl
import logging
from collections import defaultdict
from ..models import HardcodedDefault, LearnedMapping
from .excel_processor import ExcelProcessor
from .rule_engine import RuleEngine
from .validation_service import ValidationService
from .learning_engine import LearningEngine

logger = logging.getLogger(__name__)

def is_brand_attribute(attr_name: str) -> bool:
    name_lower = attr_name.lower()
    if "brand" not in name_lower:
        return False
    if name_lower.endswith(".value") or name_lower == "brand_name" or name_lower == "brand":
        if "relationship" not in name_lower and "group" not in name_lower:
            return True
    return False

def is_epi_entity_attribute(attr_name: str) -> bool:
    name_lower = attr_name.lower()
    return "external_product_information" in name_lower and name_lower.endswith(".entity")

def is_epi_value_attribute(attr_name: str) -> bool:
    name_lower = attr_name.lower()
    return "external_product_information" in name_lower and name_lower.endswith(".value")

def is_color_map_attribute(attr_name: str) -> bool:
    name_lower = attr_name.lower()
    if "color" not in name_lower:
        return False
    return "standardized_values" in name_lower or "color_map" in name_lower

def is_color_name_attribute(attr_name: str) -> bool:
    name_lower = attr_name.lower()
    if "color" not in name_lower:
        return False
    return name_lower.endswith(".value") and "standardized_values" not in name_lower and "color_map" not in name_lower

class GenerationService:
    @staticmethod
    def extract_style_code(sku):
        if not sku:
            return ""
        sku_str = str(sku).strip()
        # Common convention: StyleCode-Color-Size. Let's split by '-' or '_'
        # e.g., PUCPPA003204-DK. GREEN -> PUCPPA003204
        # e.g., PGTOPS002848-LT. GREEN -> PGTOPS002848
        parts = re_split = [x for x in sku_str.replace("_", "-").split("-") if x]
        if parts:
            return parts[0]
        return sku_str

    @classmethod
    def generate_listings(cls, db, skus_input: str, item_directory_path: str, master_sheet_path: str, content_sheet_path: str, template_path: str, output_path: str, task_logger=None):
        """
        Main runner function to generate Amazon listing flat files.
        """
        def log(msg):
            if task_logger:
                task_logger(msg)
            else:
                logger.info(msg)

        log("📂 Reading Amazon template file...")
        meta = ExcelProcessor.parse_template(template_path)
        attributes_meta = meta["attributes"]
        ptd_mappings = meta["ptd_mappings"]
        sheet_info = meta["sheet_info"]
        
        log("📊 Loading uploaded source files...")
        item_dir = LearningEngine.load_excel_as_dicts(item_directory_path) if item_directory_path else []
        master_sheet = LearningEngine.load_excel_as_dicts(master_sheet_path) if master_sheet_path else []
        content_sheet = LearningEngine.load_excel_as_dicts(content_sheet_path) if content_sheet_path else []
        
        log(f"   Loaded {len(item_dir):,} products from Item Directory")
        log(f"   Loaded {len(content_sheet):,} rows from Content Sheet")
        
        if not item_dir:
            raise ValueError("Item Directory has no rows or could not be parsed.")
            
        # Parse SKU input
        input_tokens = [x.strip() for x in skus_input.replace("\n", ",").split(",") if x.strip()]
        log(f"🔍 Searching for {len(input_tokens)} style code(s): {', '.join(input_tokens)}")
        
        # 1. Match and resolve child rows from source sheets
        # Find which column in item_dir is the SKU/Item Code column
        sku_col = None
        style_col = None
        sample_row = item_dir[0]

        # --- Step 0: Database-learned mappings lookup ---
        from ..models import LearnedMapping
        # Look up SKU columns in order of preference
        sku_attrs = [
            "contribution_sku#1.value",
            "amzn1.volt.ca.product_id_value",
            "part_number[marketplace_id=A21TJRUUN4KGV]#1.value"
        ]
        for attr in sku_attrs:
            m_sku = db.query(LearnedMapping).filter(
                LearnedMapping.amazon_attribute == attr,
                LearnedMapping.is_active == True
            ).first()
            if m_sku:
                internal_col = m_sku.internal_column
                col_key = internal_col.split(".", 1)[1] if "." in internal_col else internal_col
                if col_key in sample_row:
                    sku_col = col_key
                    break

        m_style = db.query(LearnedMapping).filter(
            LearnedMapping.amazon_attribute.like("%model_number%"),
            LearnedMapping.is_active == True
        ).first()
        if m_style:
            internal_col = m_style.internal_column
            col_key = internal_col.split(".", 1)[1] if "." in internal_col else internal_col
            if col_key in sample_row:
                style_col = col_key

        # --- Step 1: SKU Column Name-based fallback (Highly reliable standard names) ---
        if not sku_col:
            for k in sample_row.keys():
                k_clean = str(k).strip().lower()
                if k_clean in ["sku", "item_code", "item code", "itemcode"]:
                    sku_col = k
                    break

        # --- Step 2: SKU Column Format-based fallback ---
        # If still no sku_col, find a column whose values look like SKU codes (letters+digits with dash)
        if not sku_col:
            import re
            sku_pattern = re.compile(r'^[A-Z]{2,}[0-9]{4,}-[A-Z]', re.IGNORECASE)
            for k in sample_row.keys():
                v = sample_row.get(k)
                if v and sku_pattern.match(str(v).strip()):
                    sku_col = k
                    break

        # --- Step 3: SKU Column Token-driven discovery (stratified sample fallback) ---
        # Sample rows from start, middle and end to cover the full file.
        # This finds which column ACTUALLY contains values matching user's input tokens.
        if not sku_col and input_tokens:
            total = len(item_dir)
            # Build a stratified sample of up to 2000 rows spread across the file
            step = max(1, total // 2000)
            scan_rows = item_dir[::step][:2000]

            col_hit_counts: dict = {}
            for token in input_tokens:
                token_lower = token.strip().lower()
                # Extract style-code prefix: PBTSHS002794 from PBTSHS002794-BLACK
                token_prefix = token_lower.split("-")[0]
                for row in scan_rows:
                    for col_k, col_v in row.items():
                        if col_v is not None:
                            v_str = str(col_v).strip().lower()
                            if token_lower == v_str or v_str.startswith(token_prefix):
                                col_hit_counts[col_k] = col_hit_counts.get(col_k, 0) + 1

            if col_hit_counts:
                best_col = max(col_hit_counts, key=lambda c: col_hit_counts[c])
                sku_col = best_col

        # Final default SKU column fallback
        if not sku_col:
            sku_col = list(sample_row.keys())[0]

        # --- Step A: Style Code header-name exact match ---
        for k in sample_row.keys():
            k_clean = str(k).strip().lower()
            if k_clean in ["style_code", "style code", "style", "style_no", "style no", "article", "article_no", "article no"]:
                style_col = k
                break

        # --- Step B: Style Code name-based fallback (item name, style group, etc.) ---
        if not style_col:
            for k in sample_row.keys():
                k_clean = str(k).strip().lower()
                if k_clean in ["item name", "item_name", "style group", "style_group", "stylegroup"]:
                    style_col = k
                    break

        # --- Step C: Style Code pattern-based fallback ---
        if not style_col:
            for k in sample_row.keys():
                k_clean = str(k).strip().lower()
                if any(x in k_clean for x in ["style_code", "style code", "style_no", "article"]):
                    if not any(noise in k_clean for noise in ["group", "name", "category", "desc", "brand"]):
                        style_col = k
                        break

        # Column detection complete (internal)
        
        matched_child_items = []
        for token in input_tokens:
            token_str = str(token).strip().lower()
            token_prefix = token_str.split("-")[0]
            matches = []
            for row in item_dir:
                sku_val = str(row.get(sku_col, "")).strip().lower()
                style_val = str(row.get(style_col, "")).strip().lower() if style_col else ""
                
                # Check for exact or prefix matches in guessed columns
                if (token_str == sku_val or token_str == style_val or
                        sku_val.startswith(token_prefix) or style_val.startswith(token_prefix)):
                    matches.append(row)
                else:
                    # Fallback: search all columns
                    found = False
                    for col_name, val in row.items():
                        if val is not None:
                            val_str = str(val).strip().lower()
                            if token_str == val_str or val_str.startswith(token_prefix):
                                found = True
                                break
                    if found:
                        matches.append(row)
                        
            if matches:
                matched_child_items.extend(matches)
            else:
                log(f"⚠ No products found matching '{token}' — check spelling or Item Directory.")
                
        # Remove duplicates from matched children
        unique_children = []
        seen_child_skus = set()
        for c in matched_child_items:
            sku_val = c.get(sku_col)
            if sku_val not in seen_child_skus:
                seen_child_skus.add(sku_val)
                unique_children.append(c)
                
        log(f"✔ Found {len(unique_children)} matching product variant(s) across all style codes.")
        if not unique_children:
            raise ValueError("No matching products found in Item Directory for the inputted SKUs.")

        # Get learned SKU column if available
        from ..models import LearnedMapping
        learned_sku_col = None
        m_sku = db.query(LearnedMapping).filter(LearnedMapping.amazon_attribute == "contribution_sku#1.value", LearnedMapping.is_active == True).first()
        if m_sku:
            learned_sku_col = m_sku.internal_column

        # 2. Enrich child rows with joins from Master & Content sheets
        # Pre-build indices to speed up joins from O(N*M*C) to O(M*C + N)
        master_index = {}
        for m_row in master_sheet:
            for mk, mv in m_row.items():
                if mv is not None:
                    mv_str = str(mv).strip().lower()
                    if mv_str and mv_str not in master_index:
                        master_index[mv_str] = m_row

        content_index = {}
        for ct_row in content_sheet:
            for ck, cv in ct_row.items():
                if cv is not None:
                    cv_str = str(cv).strip().lower()
                    if cv_str and cv_str not in content_index:
                        content_index[cv_str] = ct_row

        joined_items = []
        for c_row in unique_children:
            sku_val = str(c_row.get(sku_col)).strip()
            sku_val_lower = sku_val.lower()
            
            # Extract parent style code smartly
            sku_for_style = sku_val
            if learned_sku_col and c_row.get(learned_sku_col) is not None:
                sku_for_style = str(c_row.get(learned_sku_col)).strip()
                
            style_val = None
            if style_col:
                val = c_row.get(style_col)
                if val is not None and str(val).strip().lower() not in ["", "(nil)", "nil", "n/a", "nan"]:
                    style_val = str(val).strip()
            
            if not style_val:
                extracted = cls.extract_style_code(sku_for_style)
                # Look for another column holding this exact value (like ITEM NAME holding style)
                found_col = None
                for k, v in c_row.items():
                    if v is not None and str(v).strip().lower() == extracted.lower():
                        found_col = k
                        break
                if found_col:
                    style_val = str(c_row.get(found_col)).strip()
                else:
                    style_val = extracted
            
            # Fallback if style_val resolved to EAN barcode or matches sku_val
            is_barcode = style_val.isdigit() and len(style_val) in [12, 13, 14]
            if style_val == sku_val or is_barcode:
                best_style = None
                for k, v in c_row.items():
                    if v is not None and "-" in str(v):
                        v_str = str(v).strip()
                        parts = [x for x in v_str.replace("_", "-").split("-") if x]
                        if parts:
                            possible_style = parts[0].strip()
                            # Check if this possible_style matches any other column's value exactly
                            matched_any = False
                            for k2, v2 in c_row.items():
                                if v2 is not None and str(v2).strip().lower() == possible_style.lower():
                                    style_val = str(v2).strip()
                                    matched_any = True
                                    break
                            if matched_any:
                                best_style = style_val
                                break
                            elif not best_style:
                                if " " not in possible_style:
                                    best_style = possible_style
                if best_style:
                    style_val = best_style
            
            style_val_lower = style_val.lower() if style_val else None
            
            # Combine child attributes
            flat_item = dict(c_row)
            
            # Join with Master Sheet
            m_matched = master_index.get(sku_val_lower) or (master_index.get(style_val_lower) if style_val_lower else None)
            if m_matched:
                for k, v in m_matched.items():
                    if k not in flat_item or flat_item[k] is None:
                        flat_item[k] = v
                        
            # Join with Content Sheet
            c_matched = content_index.get(sku_val_lower) or (content_index.get(style_val_lower) if style_val_lower else None)
            if c_matched:
                for k, v in c_matched.items():
                    if k not in flat_item or flat_item[k] is None:
                        flat_item[k] = v
                        
            joined_items.append((sku_val, style_val, flat_item))

        # 3. Group by Style Code to establish Parent-Child structures
        style_groups = defaultdict(list)
        for sku, style, flat_data in joined_items:
            style_groups[style].append(flat_data)
            
        # 4. Generate Rows for the Amazon Template
        generated_rows = []
        
        # Pre-fetch all rules, defaults, and mappings to avoid N+1 database queries
        from ..models import AdminRule, HardcodedDefault, LearnedMapping, ValueMapping
        admin_rules = db.query(AdminRule).all()
        hardcoded_defaults = db.query(HardcodedDefault).all()
        learned_mappings = db.query(LearnedMapping).all()
        value_mappings = db.query(ValueMapping).all()
        
        # Group admin rules by (amazon_attribute, scope, scope_value)
        rules_cache = {}
        for r in admin_rules:
            key = (r.amazon_attribute, r.scope, r.scope_value)
            rules_cache[key] = r
            
        # Index active defaults by amazon_attribute
        defaults_cache = {}
        for d in hardcoded_defaults:
            if d.is_active:
                defaults_cache[d.amazon_attribute] = d
                
        # Index active mappings by amazon_attribute
        mappings_cache = {}
        for m in learned_mappings:
            if m.is_active:
                mappings_cache[m.amazon_attribute] = m
                
        # Group value mappings by (amazon_attribute, internal_value)
        value_mappings_cache = {}
        for v in value_mappings:
            key = (v.amazon_attribute, v.internal_value)
            value_mappings_cache[key] = v
            
        rules_lookup = {
            "admin_rules": rules_cache,
            "defaults": defaults_cache,
            "learned_mappings": mappings_cache,
            "value_mappings": value_mappings_cache
        }
        
        # We need to resolve the Product Type. Let's look up if there's a default in db
        resolved_ptd = "SHIRT" # default fallback
        ptd_default = defaults_cache.get("product_type#1.value")
        if ptd_default and ptd_default.is_active:
            resolved_ptd = ptd_default.default_value
            
        # If the template's AttributePTDMAP doesn't support the resolved PTD, fall back to the template's PTD
        if ptd_mappings:
            if resolved_ptd not in ptd_mappings:
                available_ptds = list(ptd_mappings.keys())
                if available_ptds:
                    matched_ptd = None
                    for aptd in available_ptds:
                        if aptd.lower() in str(template_path).lower() or str(template_path).lower() in aptd.lower():
                            matched_ptd = aptd
                            break
                    resolved_ptd = matched_ptd or available_ptds[0]
                    pass  # Fallback PTD resolved internally
            
        # Get list of unlocked attributes for this Product Type
        unlocked_attrs = ptd_mappings.get(resolved_ptd, [])
        log(f"🏷 Product Type: {resolved_ptd} — {len(unlocked_attrs)} fields unlocked for this template.")
        
        for style_code, children_data in style_groups.items():
            # Tracking variables for user-friendly summary logging
            resolved_title = None
            resolved_desc = None
            bullet_points_count = 0
            child_variants_info = []
            
            # A. Generate Parent Row
            parent_sku = f"{style_code}-$P"
            parent_row = {}
            
            # VERY IMPORTANT AMAZON RULE:
            # Step 1: Set Product Type & Listing Action first
            parent_row["product_type#1.value"] = resolved_ptd
            parent_row["::record_action"] = "Create or Replace (Full Update)"
            parent_row["contribution_sku#1.value"] = parent_sku
            
            # Set parent relationship fields
            # Find exact technical name for parentage level & variation theme
            parentage_attr = None
            theme_attr = None
            parent_sku_attr = None
            relationship_type_attr = None
            
            for attr in attributes_meta.keys():
                if "parentage_level" in attr:
                    parentage_attr = attr
                if "variation_theme#1.name" in attr or "variation_theme" in attr:
                    theme_attr = attr
                if "parent_sku" in attr:
                    parent_sku_attr = attr
                if "relationship_type" in attr:
                    relationship_type_attr = attr
            
            if parentage_attr:
                parent_row[parentage_attr] = "Parent"
            if theme_attr:
                parent_row[theme_attr] = "SIZE/COLOR"
                
            # For parent, populate other applicable fields from the first child's data
            sample_child = children_data[0]
            
            for attr, attr_info in attributes_meta.items():
                # Skip basic keys already set
                if attr in ["product_type#1.value", "::record_action", "contribution_sku#1.value", parentage_attr, theme_attr, parent_sku_attr, relationship_type_attr]:
                    continue
                    
                # Skip attributes that are locked for this PTD (if conditional attribute)
                is_conditional = any(attr in attrs for attrs in ptd_mappings.values())
                if is_conditional and attr not in unlocked_attrs:
                    continue
                    
                # Skip variation attributes that MUST be blank on Parent row
                is_variation_field = any(x in attr.lower() for x in ["size", "color", "price", "product_id", "external_product_information", "barcode"])
                if is_variation_field:
                    continue
                    
                # Resolve value using Rule Engine
                val, source_type, score = RuleEngine.resolve_attribute_value(
                    db, attr, sample_child, product_type=resolved_ptd, brand=sample_child.get("Brand"), category=sample_child.get("Category"), rules_lookup=rules_lookup
                )
                
                # Rule Overrides
                if is_brand_attribute(attr):
                    div_val = ""
                    for k, v in sample_child.items():
                        if k.strip().lower() == "division":
                            div_val = str(v).strip().upper() if v is not None else ""
                            break
                    if div_val == "FOOTWEAR":
                        val = "Toothless"
                        source_type = "override"
                    elif div_val in ["APPAREL", "ACCESSORIES"]:
                        val = "Purple United Kids"
                        source_type = "override"
                        
                elif is_color_map_attribute(attr):
                    if val is not None:
                        val = str(val).title()
                        source_type = "override"
                if val is not None:
                    if "bullet_point" in attr.lower():
                        import re
                        bullet_match = re.search(r'#(\d+)\.value', attr)
                        if bullet_match:
                            bullet_idx = int(bullet_match.group(1))
                            lines = [l.strip().lstrip('*').lstrip('-').lstrip('●').lstrip('•').strip() 
                                     for l in str(val).replace('\r', '\n').split('\n') if l.strip()]
                            if len(lines) >= bullet_idx:
                                val = lines[bullet_idx - 1]
                                bullet_points_count = max(bullet_points_count, bullet_idx)
                            else:
                                val = None
                    if val is not None:
                        parent_row[attr] = val
                        if "item_name" in attr.lower():
                            resolved_title = val
                        elif "product_description" in attr.lower():
                            resolved_desc = val
                    
            generated_rows.append(parent_row)
            
            # B. Generate Child Rows
            for child_idx, child_flat in enumerate(children_data):
                child_row = {}
                child_sku_val = child_flat.get(sku_col, f"{style_code}-CHILD-{child_idx}")
                
                # Step 1: Set Product Type & Listing Action first
                child_row["product_type#1.value"] = resolved_ptd
                child_row["::record_action"] = "Create or Replace (Full Update)"
                child_row["contribution_sku#1.value"] = child_sku_val
                
                if parentage_attr:
                    child_row[parentage_attr] = "Child"
                if parent_sku_attr:
                    child_row[parent_sku_attr] = parent_sku
                if theme_attr:
                    child_row[theme_attr] = "SIZE/COLOR"
                if relationship_type_attr:
                    child_row[relationship_type_attr] = "Variation"
                    
                # Populate remaining attributes
                for attr, attr_info in attributes_meta.items():
                    if attr in [
                        "product_type#1.value", "::record_action", "contribution_sku#1.value",
                        parentage_attr, parent_sku_attr, theme_attr, relationship_type_attr
                    ]:
                        continue
                        
                    # Skip locked attributes (if conditional attribute)
                    is_conditional = any(attr in attrs for attrs in ptd_mappings.values())
                    if is_conditional and attr not in unlocked_attrs:
                        continue
                        
                    # Resolve value using Rule Engine
                    val, source_type, score = RuleEngine.resolve_attribute_value(
                        db, attr, child_flat, product_type=resolved_ptd, brand=child_flat.get("Brand"), category=child_flat.get("Category"), rules_lookup=rules_lookup
                    )
                    
                    # Rule Overrides
                    if is_brand_attribute(attr):
                        div_val = ""
                        for k, v in child_flat.items():
                            if k.strip().lower() == "division":
                                div_val = str(v).strip().upper() if v is not None else ""
                                break
                        if div_val == "FOOTWEAR":
                            val = "Toothless"
                            source_type = "override"
                        elif div_val in ["APPAREL", "ACCESSORIES"]:
                            val = "Purple United Kids"
                            source_type = "override"
                            
                    elif is_epi_entity_attribute(attr):
                        val = "HSN Code"
                        source_type = "override"
                        
                    elif is_epi_value_attribute(attr):
                        hs_val = ""
                        for k, v in child_flat.items():
                            if k.strip().lower() in ["hs code", "hscode", "hsn code", "hsncode"]:
                                hs_val = str(v).strip() if v is not None else ""
                                break
                        if hs_val:
                            val = hs_val
                            source_type = "override"
                            
                    elif is_color_name_attribute(attr):
                        c_val = ""
                        for k, v in child_flat.items():
                            if k.strip().lower() in ["color", "color_name", "item_color", "item color"]:
                                c_val = str(v).strip() if v is not None else ""
                                break
                        if c_val:
                            val = c_val
                            source_type = "override"
                            
                    elif is_color_map_attribute(attr):
                        c_val = ""
                        for k, v in child_flat.items():
                            if k.strip().lower() in ["color", "color_name", "item_color", "item color"]:
                                c_val = str(v).strip() if v is not None else ""
                                break
                        if c_val:
                            val = c_val.title()
                            source_type = "override"
                    if val is not None:
                        if "bullet_point" in attr.lower():
                            import re
                            bullet_match = re.search(r'#(\d+)\.value', attr)
                            if bullet_match:
                                bullet_idx = int(bullet_match.group(1))
                                lines = [l.strip().lstrip('*').lstrip('-').lstrip('●').lstrip('•').strip() 
                                         for l in str(val).replace('\r', '\n').split('\n') if l.strip()]
                                if len(lines) >= bullet_idx:
                                    val = lines[bullet_idx - 1]
                                else:
                                    val = None
                        if val is not None:
                            child_row[attr] = val
                        
                generated_rows.append(child_row)
                
                # Track variant details for log summary
                child_size = "N/A"
                child_color = "N/A"
                for k, v in child_flat.items():
                    k_lower = k.lower()
                    if v is not None:
                        if k_lower in ["size", "footwear_size", "size_name", "item_size"]:
                            child_size = str(v)
                        elif k_lower in ["color", "color_name", "color_map"]:
                            child_color = str(v)
                child_variants_info.append({
                    "sku": child_sku_val,
                    "size": child_size,
                    "color": child_color
                })

            # Print friendly user-facing summary for this style group
            log(f"✅ Resolved & Generated Listing for Style '{style_code}':")
            if resolved_title:
                log(f"   • Product Title: \"{resolved_title}\"")
            else:
                log(f"   • Product Title: (Not found in Content Sheet)")
                
            if resolved_desc:
                desc_summary = resolved_desc[:120] + ("..." if len(resolved_desc) > 120 else "")
                log(f"   • Description: \"{desc_summary}\"")
            else:
                log(f"   • Description: (Not found in Content Sheet)")
                
            log(f"   • Bullet Points: Loaded {bullet_points_count} points")
            log(f"   • Child Variants ({len(child_variants_info)} items):")
            for c_info in child_variants_info:
                log(f"     - SKU: {c_info['sku']} (Size: {c_info['size']}, Color: {c_info['color']})")
            log("")  # Empty line for padding

        log("🔎 Validating all generated rows against Amazon requirements...")
        val_report = ValidationService.validate_listings(generated_rows, attributes_meta, ptd_mappings)
        if val_report["errors_count"] == 0:
            log(f"✅ Validation passed — {val_report['warnings_count']} warning(s).")
        else:
            log(f"⚠ Validation found {val_report['errors_count']} error(s) and {val_report['warnings_count']} warning(s).")
        
        # 5. Write to Template Sheet and Save
        log("📝 Writing data into the Amazon template file...")
        
        # Open template with keep_vba=True
        wb_write = openpyxl.load_workbook(template_path, keep_vba=True)
        sheet_write = wb_write['Template']
        
        # Start writing at data_row
        start_row = sheet_info.get("data_row", 7)
        
        # Clear existing data rows if any (up to row 500)
        for r_idx in range(start_row, max(start_row + len(generated_rows) + 50, sheet_write.max_row + 1)):
            for c_idx in range(1, sheet_write.max_column + 1):
                sheet_write.cell(row=r_idx, column=c_idx).value = None
                
        # Write new generated rows
        for r_idx, row_data in enumerate(generated_rows):
            target_row = start_row + r_idx
            for attr, val in row_data.items():
                if attr in attributes_meta:
                    col_idx = attributes_meta[attr]["column_index"]
                    # openpyxl columns are 1-indexed
                    sheet_write.cell(row=target_row, column=col_idx + 1).value = val
                    
        log(f"Saving completed template to output file: {output_path}")
        # Ensure output directory exists
        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
            
        wb_write.save(output_path)
        wb_write.close()
        log("Listing generation successfully finished!")
        
        return {
            "status": "success",
            "total_rows": len(generated_rows),
            "parent_rows": len(style_groups),
            "child_rows": len(joined_items),
            "validation": val_report,
            "output_file": output_path
        }
