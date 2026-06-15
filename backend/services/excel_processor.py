import openpyxl
from openpyxl.utils import get_column_letter
import re

class ExcelProcessor:
    @staticmethod
    def clean_label(label):
        if not label:
            return ""
        # Remove suffix like " - [ SHIRT ]" or " - [  ]"
        return re.sub(r"\s*-\s*\[.*\]\s*$", "", str(label)).strip()

    @classmethod
    def parse_template(cls, filepath):
        """
        Parses an Amazon category template and returns metadata.
        Returns:
            dict containing:
                - 'attributes': dict of tech_name -> attribute info
                - 'ptd_mappings': dict of product_type -> list of tech_names
                - 'valid_values': dict of label -> list of allowed values
                - 'sheet_info': dict of template indices and layout
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        # 1. Parse Valid Values Sheet
        valid_values_map = {}
        if 'Valid Values' in wb.sheetnames:
            sheet_vv = wb['Valid Values']
            for row in sheet_vv.iter_rows(values_only=True):
                # Check column 2 (index 1)
                if len(row) > 1 and row[1]:
                    raw_label = str(row[1])
                    if " - [" in raw_label:
                        clean_lbl = cls.clean_label(raw_label)
                        # The remaining cells in the row are valid values
                        allowed = [str(x).strip() for x in row[2:] if x is not None]
                        if allowed:
                            valid_values_map[clean_lbl] = allowed

        # 2. Parse Data Definitions Sheet
        definitions_map = {}
        if 'Data Definitions' in wb.sheetnames:
            sheet_dd = wb['Data Definitions']
            rows_dd = list(sheet_dd.iter_rows(values_only=True))
            # Find the header row (Group Name, Field Name, Local Label Name, etc.)
            header_row_idx = 1 # 0-based
            for idx, r in enumerate(rows_dd[:5]):
                if r and 'Field Name' in r and 'Required?' in r:
                    header_row_idx = idx
                    break
            
            headers = rows_dd[header_row_idx]
            field_name_idx = headers.index('Field Name') if 'Field Name' in headers else 1
            label_idx = headers.index('Local Label Name') if 'Local Label Name' in headers else 2
            desc_idx = headers.index('Accepted Values') if 'Accepted Values' in headers else 3
            required_idx = headers.index('Required?') if 'Required?' in headers else 5
            
            for r in rows_dd[header_row_idx + 1:]:
                if len(r) > field_name_idx and r[field_name_idx]:
                    tech_name = str(r[field_name_idx]).strip()
                    label_name = str(r[label_idx]).strip() if len(r) > label_idx and r[label_idx] else ""
                    required_status = str(r[required_idx]).strip() if len(r) > required_idx and r[required_idx] else "Optional"
                    desc = str(r[desc_idx]).strip() if len(r) > desc_idx and r[desc_idx] else ""
                    
                    definitions_map[tech_name] = {
                        "label": label_name,
                        "required": required_status,
                        "description": desc
                    }

        # 3. Parse Template Sheet Columns
        attributes = {}
        sheet_info = {}
        if 'Template' in wb.sheetnames:
            sheet_t = wb['Template']
            rows_t = list(sheet_t.iter_rows(max_row=10, values_only=True))
            
            # Identify settings, labels, tech_names
            # Defaults if we can't parse them from settings:
            label_row_idx = 3 # 0-based (Row 4)
            attr_row_idx = 4  # 0-based (Row 5)
            data_row_idx = 6  # 0-based (Row 7)
            
            # Row 1 often contains settings
            if rows_t and rows_t[0] and rows_t[0][0] and "settings=" in str(rows_t[0][0]):
                settings_str = str(rows_t[0][0])
                label_match = re.search(r"labelRow=(\d+)", settings_str)
                attr_match = re.search(r"attributeRow=(\d+)", settings_str)
                data_match = re.search(r"dataRow=(\d+)", settings_str)
                if label_match:
                    label_row_idx = int(label_match.group(1)) - 1
                if attr_match:
                    attr_row_idx = int(attr_match.group(1)) - 1
                if data_match:
                    data_row_idx = int(data_match.group(1)) - 1
            
            sheet_info = {
                "label_row": label_row_idx + 1,
                "attribute_row": attr_row_idx + 1,
                "data_row": data_row_idx + 1
            }
            
            labels = rows_t[label_row_idx]
            tech_names = rows_t[attr_row_idx]
            
            for col_idx in range(len(tech_names)):
                tech_name = tech_names[col_idx]
                if tech_name:
                    tech_name = str(tech_name).strip()
                    label = str(labels[col_idx]).strip() if col_idx < len(labels) and labels[col_idx] else ""
                    col_letter = get_column_letter(col_idx + 1)
                    
                    # Merge with Data Definitions
                    def_info = definitions_map.get(tech_name, {})
                    required = def_info.get("required", "Optional")
                    
                    # Fetch valid values matching by local label name or tech name
                    clean_lbl = cls.clean_label(label)
                    allowed_values = valid_values_map.get(clean_lbl, None)
                    if not allowed_values:
                        allowed_values = valid_values_map.get(label, None)
                        
                    attributes[tech_name] = {
                        "technical_name": tech_name,
                        "label": label or def_info.get("label", ""),
                        "column_letter": col_letter,
                        "column_index": col_idx,
                        "required": required,
                        "valid_values": allowed_values,
                        "description": def_info.get("description", "")
                    }

        # 4. Parse AttributePTDMAP Sheet
        ptd_mappings = {}
        if 'AttributePTDMAP' in wb.sheetnames:
            sheet_ptd = wb['AttributePTDMAP']
            rows_ptd = list(sheet_ptd.iter_rows(values_only=True))
            if rows_ptd:
                # Row 1 contains Product Types (PTDs) starting from col 2 (index 1)
                header = rows_ptd[0]
                ptds = []
                for col_idx in range(1, len(header)):
                    if header[col_idx]:
                        ptds.append((col_idx, str(header[col_idx]).strip()))
                
                # Rows 2 onwards contain attribute names in col 1 and applicability (1) in PTD columns
                for r in rows_ptd[1:]:
                    if r and r[0]:
                        attr_name = str(r[0]).strip()
                        for col_idx, ptd_name in ptds:
                            if col_idx < len(r) and (r[col_idx] == 1 or str(r[col_idx]) == '1'):
                                if ptd_name not in ptd_mappings:
                                    ptd_mappings[ptd_name] = []
                                ptd_mappings[ptd_name].append(attr_name)
                                
        wb.close()
        return {
            "attributes": attributes,
            "ptd_mappings": ptd_mappings,
            "valid_values": valid_values_map,
            "sheet_info": sheet_info
        }
