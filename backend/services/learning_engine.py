import pandas as pd
import openpyxl
from collections import defaultdict
from sqlalchemy.orm import Session
from ..models import LearnedMapping, ValueMapping, HardcodedDefault
from .excel_processor import ExcelProcessor

class LearningEngine:
    @staticmethod
    def load_excel_as_dicts(filepath):
        """Loads an excel sheet as a list of dictionaries (keys are header names)."""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # If it has a 'Template' sheet, use it (Amazon template style)
        if 'Template' in wb.sheetnames:
            sheetname_to_use = 'Template'
        else:
            sheetname_to_use = None
            for name in wb.sheetnames:
                sheet = wb[name]
                if sheet.max_row > 1:
                    sheetname_to_use = name
                    break
        
        if not sheetname_to_use:
            return []
            
        sheet = wb[sheetname_to_use]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
            
        # Find first row with at least some strings to treat as headers
        header_idx = 0
        if sheetname_to_use == 'Template':
            # For Amazon templates, check if Row 1 has settings, then row 5 contains attribute names (index 4)
            if rows[0] and rows[0][0] and "settings=" in str(rows[0][0]):
                header_idx = 4
            else:
                for idx, r in enumerate(rows[:5]):
                    if any(isinstance(x, str) for x in r if x):
                        header_idx = idx
                        break
        else:
            for idx, r in enumerate(rows[:5]):
                if any(isinstance(x, str) for x in r if x):
                    header_idx = idx
                    break
                    
        raw_headers = [str(h).strip() if h is not None else f"Col_{i}" for i, h in enumerate(rows[header_idx])]
        headers = []
        seen = {}
        for h in raw_headers:
            if h in seen:
                seen[h] += 1
                headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                headers.append(h)
        data = []
        
        # Data rows start after header. For Amazon templates, start at Row 7 (index 6)
        start_row_idx = header_idx + 2 if (sheetname_to_use == 'Template' and header_idx == 4) else header_idx + 1
        
        for r in rows[start_row_idx:]:
            if any(x is not None for x in r):
                row_dict = {}
                for i, val in enumerate(r):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                data.append(row_dict)
        wb.close()
        return data

    @classmethod
    def learn_from_historical_listing(cls, db: Session, history_filepath: str, item_directory_path: str = None, master_sheet_path: str = None, content_sheet_path: str = None):
        """
        Analyzes a completed Amazon listing Excel file and matches it against source sheets to learn mappings.
        """
        # 1. Parse historical Amazon listing using ExcelProcessor
        template_meta = ExcelProcessor.parse_template(history_filepath)
        attributes = template_meta["attributes"]
        
        # Load data rows from history template (starts from data_row row)
        wb_hist = openpyxl.load_workbook(history_filepath, data_only=True)
        sheet_t = wb_hist['Template']
        
        sheet_info = template_meta["sheet_info"]
        data_start_row = sheet_info.get("data_row", 7)
        attr_row = sheet_info.get("attribute_row", 5)
        
        tech_names = [cell.value for cell in sheet_t[attr_row]]
        
        history_rows = []
        for r_idx in range(data_start_row, sheet_t.max_row + 1):
            row_vals = [sheet_t.cell(row=r_idx, column=c_idx + 1).value for c_idx in range(len(tech_names))]
            if any(v is not None for v in row_vals):
                row_dict = {tech_names[i]: row_vals[i] for i in range(len(tech_names)) if tech_names[i]}
                history_rows.append(row_dict)
                
        if not history_rows:
            wb_hist.close()
            return {"status": "error", "message": "No data rows found in historical template"}
            
        # 2. Load source files
        sources = {}
        if item_directory_path:
            sources["item_directory"] = cls.load_excel_as_dicts(item_directory_path)
        if master_sheet_path:
            sources["master_sheet"] = cls.load_excel_as_dicts(master_sheet_path)
        if content_sheet_path:
            sources["content_sheet"] = cls.load_excel_as_dicts(content_sheet_path)
            
        # 3. Detect Join Columns
        # We need to find which column in each source sheet corresponds to the SKU / Product ID in the history template
        # We look at historical SKUs (contribution_sku#1.value) and Product ID (amzn1.volt.ca.product_id_value)
        hist_skus = set(str(row.get("contribution_sku#1.value")).strip() for row in history_rows if row.get("contribution_sku#1.value"))
        hist_pids = set(str(row.get("amzn1.volt.ca.product_id_value")).strip() for row in history_rows if row.get("amzn1.volt.ca.product_id_value"))
        
        join_columns = {}
        for src_name, src_rows in sources.items():
            if not src_rows:
                continue
            best_col = None
            max_overlap = 0
            # Test each column in the source sheet
            cols = list(src_rows[0].keys())
            for col in cols:
                col_vals = set(str(r.get(col)).strip() for r in src_rows if r.get(col) is not None)
                overlap_sku = len(col_vals.intersection(hist_skus))
                overlap_pid = len(col_vals.intersection(hist_pids))
                overlap = max(overlap_sku, overlap_pid)
                if overlap > max_overlap:
                    max_overlap = overlap
                    best_col = col
            if best_col and max_overlap > 0:
                join_columns[src_name] = best_col

        # 4. Perform Matching and Learn Mappings
        # Map Amazon Attribute -> Source File Column
        mapping_candidates = defaultdict(lambda: defaultdict(int))
        # Map Amazon Attribute -> Value Pairs for translation learning
        value_pairs = defaultdict(lambda: defaultdict(int))
        # Constant Detection: Map Amazon Attribute -> Constant Values count
        attr_values_count = defaultdict(lambda: defaultdict(int))
        
        matched_count = 0
        
        # Hardcoded shipping / safety compliance prefixes to NOT map to columns
        hardcoded_prefixes = [
            "item_dimensions", "item_package_dimensions", "item_package_weight",
            "country_of_origin", "item_weight", "packer_contact_information",
            "rtip_manufacturer_contact_information", "safety_and_compliance",
            "regulatory_compliance", "gpsr_safety", "compliance_", "importer_contact_information"
        ]
        
        # Pre-index source rows for O(1) lookups
        dir_rows = sources.get("item_directory", [])
        dir_join_col = join_columns.get("item_directory")
        dir_by_join = {}
        if dir_rows and dir_join_col:
            for s_row in dir_rows:
                s_val = str(s_row.get(dir_join_col)).strip()
                if s_val:
                    dir_by_join[s_val] = s_row
                    
        mst_rows = sources.get("master_sheet", [])
        mst_join_col = join_columns.get("master_sheet")
        mst_by_join = {}
        if mst_rows and mst_join_col:
            for s_row in mst_rows:
                s_val = str(s_row.get(mst_join_col)).strip()
                if s_val:
                    mst_by_join[s_val] = s_row
                    
        cnt_rows = sources.get("content_sheet", [])
        cnt_join_col = join_columns.get("content_sheet")
        cnt_by_join = {}
        cnt_by_style = {}
        if cnt_rows:
            for s_row in cnt_rows:
                if cnt_join_col:
                    s_val = str(s_row.get(cnt_join_col)).strip()
                    if s_val:
                        cnt_by_join[s_val] = s_row
                for ck in ["Item Name", "item_name", "SKU", "sku"]:
                    cv = s_row.get(ck)
                    if cv:
                        cv_str = str(cv).strip()
                        cnt_by_style[cv_str] = s_row
                        parts = cv_str.split("-")
                        if parts:
                            cnt_by_style[parts[0]] = s_row

        for h_row in history_rows:
            sku = str(h_row.get("contribution_sku#1.value")).strip()
            pid = str(h_row.get("amzn1.volt.ca.product_id_value")).strip()
            
            # Find matching rows in sources
            matched_src_rows = {}
            
            # 4a. Match item_directory
            dir_match = None
            if dir_by_join:
                dir_match = dir_by_join.get(sku) or dir_by_join.get(pid)
                if dir_match:
                    matched_src_rows["item_directory"] = dir_match
                    
            # 4b. Match master_sheet
            mst_match = None
            if mst_by_join:
                mst_match = mst_by_join.get(sku) or mst_by_join.get(pid)
                if mst_match:
                    matched_src_rows["master_sheet"] = mst_match
                    
            # 4c. Match content_sheet
            cnt_match = None
            if cnt_by_join:
                cnt_match = cnt_by_join.get(sku) or cnt_by_join.get(pid)
                
            if not cnt_match and cnt_by_style:
                # Try relational match using style code from item_directory or master_sheet
                ref_row = dir_match or mst_match
                if ref_row:
                    style_code = None
                    for k in ["ITEM NAME", "item_name", "STYLE GROUP", "Style", "Article", "ARTICLE"]:
                        if ref_row.get(k):
                            style_code = str(ref_row.get(k)).strip()
                            break
                    if not style_code:
                        # Extract style prefix from first key containing a dash
                        for k, v in ref_row.items():
                            if v and "-" in str(v):
                                v_str = str(v).strip()
                                parts = [x for x in v_str.replace("_", "-").split("-") if x]
                                if parts and len(parts[0]) >= 4:
                                    style_code = parts[0]
                                    break
                                    
                    if style_code:
                        cnt_match = cnt_by_style.get(style_code)
                        
            if cnt_match:
                matched_src_rows["content_sheet"] = cnt_match
            
            if not matched_src_rows:
                # If we couldn't match this row, still track constant values
                for attr, h_val in h_row.items():
                    if h_val is not None:
                        attr_values_count[attr][str(h_val).strip()] += 1
                continue
                
            matched_count += 1
            
            # Compare values
            for attr, h_val in h_row.items():
                if h_val is None:
                    continue
                h_val_str = str(h_val).strip()
                attr_values_count[attr][h_val_str] += 1
                
                # If this attribute belongs to hardcoded categories, do not map to columns
                if any(attr.startswith(p) for p in hardcoded_prefixes):
                    continue
                
                # Check each source sheet column value
                for src_name, s_row in matched_src_rows.items():
                    for s_col, s_val in s_row.items():
                        if s_val is None:
                            continue
                        s_val_str = str(s_val).strip()
                        s_val_str_lower = s_val_str.lower()
                        h_val_str_lower = h_val_str.lower()
                        
                        # Direct equivalence (case-insensitive) - weighted higher (+2)
                        if s_val_str_lower == h_val_str_lower:
                            mapping_candidates[attr][f"{src_name}.{s_col}"] += 2
                        # Substring equivalence for long values (case-insensitive) - weighted (+1)
                        elif len(h_val_str) > 10 and h_val_str_lower in s_val_str_lower:
                            mapping_candidates[attr][f"{src_name}.{s_col}"] += 1
                        
                        # Track for translation mapping (e.g. size/color mappings)
                        if ("size" in s_col.lower() and "size" in attr.lower()) or ("color" in s_col.lower() and "color" in attr.lower()):
                            value_pairs[attr][f"{s_val_str}||{h_val_str}"] += 1

        # 5. Save Learned Mappings to Database
        learned_mappings_created = 0
        value_mappings_created = 0
        defaults_created = 0
        
        # Min confidence threshold (at least 30% match rate of matched rows)
        min_matches = max(1, int(matched_count * 0.3))
        # Value mappings threshold is lower (since distinct sizes/colors appear in subsets of rows)
        min_value_matches = 2
        
        # Save mappings
        for attr, candidates in mapping_candidates.items():
            best_candidate = None
            max_votes = 0
            for cand, votes in candidates.items():
                if votes > max_votes and votes >= min_matches:
                    max_votes = votes
                    best_candidate = cand
            
            if best_candidate:
                confidence = max_votes / (matched_count * 2.0)
                # Save/Update in db
                mapping_obj = db.query(LearnedMapping).filter(LearnedMapping.amazon_attribute == attr).first()
                if not mapping_obj:
                    mapping_obj = LearnedMapping(amazon_attribute=attr)
                    db.add(mapping_obj)
                mapping_obj.internal_column = best_candidate
                mapping_obj.confidence_score = confidence
                mapping_obj.is_active = True
                learned_mappings_created += 1
                
        # Save Value Translations
        for attr, pairs in value_pairs.items():
            for pair, count in pairs.items():
                if count >= min_value_matches:
                    internal_val, amazon_val = pair.split("||", 1)
                    confidence = count / matched_count
                    val_obj = db.query(ValueMapping).filter(
                        ValueMapping.amazon_attribute == attr,
                        ValueMapping.internal_value == internal_val
                    ).first()
                    if not val_obj:
                        val_obj = ValueMapping(amazon_attribute=attr, internal_value=internal_val)
                        db.add(val_obj)
                    val_obj.amazon_value = amazon_val
                    val_obj.confidence_score = confidence
                    value_mappings_created += 1
                    
        # Save Defaults (Constant attributes)
        # An attribute is a default if it has the same value for ALL rows (>95% of rows)
        # AND it does not map to any source column
        for attr, val_counts in attr_values_count.items():
            # Check if this attribute was mapped to a source column
            is_mapped = db.query(LearnedMapping).filter(LearnedMapping.amazon_attribute == attr, LearnedMapping.is_active == True).first() is not None
            if is_mapped:
                continue
                
            for val, count in val_counts.items():
                if count >= int(len(history_rows) * 0.95):
                    # We have a default!
                    def_obj = db.query(HardcodedDefault).filter(HardcodedDefault.amazon_attribute == attr).first()
                    if not def_obj:
                        def_obj = HardcodedDefault(amazon_attribute=attr)
                        db.add(def_obj)
                    def_obj.default_value = val
                    def_obj.is_active = True
                    defaults_created += 1
                    break
                    
        wb_hist.close()
        db.commit()
        return {
            "status": "success",
            "matched_rows": matched_count,
            "total_historical_rows": len(history_rows),
            "learned_column_mappings": learned_mappings_created,
            "learned_value_mappings": value_mappings_created,
            "detected_defaults": defaults_created
        }
